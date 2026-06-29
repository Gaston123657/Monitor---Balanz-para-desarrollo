"""Rellena SOLO metadata segura (short_name + sector) de las ONs que LSEG no
pudo resolver inequívocamente, infiriéndola del ticker HERMANO (mismo emisor,
prefijo de root compartido). NO toca isin/fechas/cupón/legislacion/base, que
dependen de la serie concreta y se cargan cuando se confirme el ISIN.

    py -3.12 scripts/apply_ons_sibling_meta.py [--apply]
"""
import argparse
import os
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
XLSX = str(ROOT / "data" / "instruments_master.xlsx")

# Tickers que quedaron sin resolver (ambiguos en LSEG o sin data).
TARGETS = ["BACHD", "CP38D", "DNCAD", "GN49D", "VSCXD", "EAC4D", "OZC8D",
           "MR46D", "MR47D", "YFCDD", "CLI1D", "CLSID", "NBS1D"]


def main():
    ap = argparse.ArgumentParser(); ap.add_argument("--apply", action="store_true")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb["ONs"]; rows = list(ws.iter_rows(values_only=True))
    hdr = [str(c).strip().lower() if c else "" for c in rows[0]]
    N, S = hdr.index("short_name"), hdr.index("sector")
    m = {}
    for r in rows[1:]:
        if r and r[0]:
            m[str(r[0]).upper()] = (r[N], r[S])
    wb.close()

    def sibling(tk):
        best = None
        for k, v in m.items():
            if k == tk or not v[0] or not v[1]:
                continue
            cp = len(os.path.commonprefix([k, tk]))
            if cp >= 3 and (best is None or cp > best[0]):
                best = (cp, k, v)
        return best

    plan = []
    for tk in TARGETS:
        cur_name, cur_sec = m.get(tk, (None, None))
        sib = sibling(tk)
        f = {"ticker": tk}
        if sib:
            _, _, (nm, sec) = sib
            if not cur_name and nm:
                f["short_name"] = nm
            if not cur_sec and sec:
                f["sector"] = sec
        if len(f) > 1:
            plan.append((tk, f, sib[1] if sib else None))

    print(f"{'APLICANDO' if args.apply else 'DRY-RUN'} · {len(plan)} ONs (metadata)\n")
    for tk, f, sib in plan:
        print(f"  {tk:<7} <- sib {sib:<7} {{ {', '.join(f'{k}={v!r}' for k, v in f.items() if k != 'ticker')} }}")
    no_sib = [tk for tk in TARGETS if tk not in {p[0] for p in plan}]
    if no_sib:
        print(f"\nSin hermano (no se toca, requiere web/manual): {no_sib}")

    if args.apply:
        from apps.web.instruments_abm import _LOCK, _write_instrument_row_on_wb, _atomic_save_workbook
        with _LOCK:
            wbw = openpyxl.load_workbook(XLSX)
            try:
                for tk, f, _ in plan:
                    _write_instrument_row_on_wb(wbw, "ONs", tk, f)
                _atomic_save_workbook(wbw, XLSX)
            finally:
                wbw.close()
        print(f"\nEscritas: {len(plan)} ONs")


if __name__ == "__main__":
    main()
