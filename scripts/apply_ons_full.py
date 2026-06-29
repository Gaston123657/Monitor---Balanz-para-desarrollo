"""Paso 3/3 (full): completa los campos faltantes de la hoja ONs del master.

    py -3.12 scripts/apply_ons_full.py            # dry-run (no escribe, imprime plan)
    py -3.12 scripts/apply_ons_full.py --apply    # escribe el Excel (1 transacción)

Lee data/lseg_ons_full.json (de scripts/lseg_fetch_ons_full.py) + los valores
actuales de cada fila, y rellena SOLO LOS CAMPOS EN BLANCO (nunca pisa lo que
el operador ya cargó). Derivaciones:
  - short_name, isin, fecha_emision, fecha_vencimiento, cupón, frecuencia: de LSEG.
  - legislacion: del prefijo del ISIN (US*->NY, AR*->AR). Regla validada contra
    las filas ya cargadas (US->NY 31/32, AR->AR 107/107).
  - base calculo: de la legislacion (NY->30/360, AR->ACT/365). Validado contra
    las filas cargadas (NY 31/31, AR 110/119 dominante).
  - sector: mapeo issuer-name + TRBC -> taxonomía Balanz. Se imprimen TODOS en
    el dry-run para revisión; los no mapeables se listan para carga manual/web.
  - tipo amortizacion + amort inicio/cantidad/frec/%: del schedule IPA.

NO toca la hoja Cashflows: las flows exactas ya están cargadas para las ONs
resueltas antes, y para las nuevas el synth (_synth_on_bond) las genera de los
campos que completamos acá.

Escribe SIEMPRE vía instruments_abm (único escritor sancionado del Excel).
"""
import argparse
import json
import re
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from apps.web import instruments_abm  # noqa: E402

XLSX = str(ROOT / "data" / "instruments_master.xlsx")
FULL = ROOT / "data" / "lseg_ons_full.json"
STD_FREQ = {1, 2, 4, 12}

# Sufijos legales / ruido que se sacan para normalizar el nombre del emisor
# y matchear contra la clasificación que el operador YA cargó en otras filas.
_LEGAL = re.compile(
    r"\b(S\.?A\.?C?\.?I?\.?F?\.?I?\.?A?|SAU|SRL|SAIC|SAAI|SACIFIA|SACIF|"
    r"S A U|Y A|LTDA|SL|BRANCH|COMPANIA FINANCIERA|ARGENTINA|BUENOS AIRES|SA)\b")


def _norm_issuer(s):
    if not s:
        return ""
    s = re.sub(r"[^A-Z0-9 ]", " ", str(s).upper())
    s = _LEGAL.sub(" ", s)
    return re.sub(r"\s+", " ", s).strip()


def load_operator_sector_map(xlsx):
    """{issuer_normalizado: sector} a partir de las filas ONs YA clasificadas
    por el operador. Es la fuente PRIMARIA de sector (respeta su taxonomía)."""
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    try:
        ws = wb["ONs"]
        rows = list(ws.iter_rows(values_only=True))
        hdr = [str(c).strip().lower() if c else "" for c in rows[0]]
        ni, si = hdr.index("short_name"), hdr.index("sector")
        votes = defaultdict(lambda: defaultdict(int))
        for r in rows[1:]:
            if not r or not r[0]:
                continue
            nm, sec = r[ni], r[si]
            if nm and sec:
                k = _norm_issuer(nm)
                if k:
                    votes[k][str(sec).strip()] += 1
        # por emisor, el sector más frecuente (desempata IRSA SA vs IRSA Inv.)
        return {k: max(v, key=v.get) for k, v in votes.items()}
    finally:
        wb.close()

BALANZ_SECTORS = {
    "Petróleo y Gas", "Servicios Públicos y Generación", "Sector Financiero",
    "Telecomunicaciones", "Consumo Masivo y Agro", "Construcción",
    "Bienes Raíces", "Infraestructura y Transporte",
}


def _to_date(iso):
    if not iso:
        return None
    return datetime.strptime(str(iso)[:10], "%Y-%m-%d").date()


def _num(v):
    if v in (None, ""):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def derive_legislacion(isin):
    if not isin:
        return None
    p = isin[:2].upper()
    if p == "US":          # USP.../US... = ley extranjera (NY)
        return "NY"
    if p == "AR":
        return "AR"
    return None


def derive_base_calculo(leg):
    return {"NY": "30/360", "AR": "ACT/365"}.get(leg)


# TRBC Economic Sector -> taxonomía Balanz (fallback final).
_TRBC_MAP = {
    "Financials": "Sector Financiero",
    "Energy": "Petróleo y Gas",
    "Utilities": "Servicios Públicos y Generación",
    "Real Estate": "Bienes Raíces",
    "Telecommunications Services": "Telecomunicaciones",
    "Consumer Non-Cyclicals": "Consumo Masivo y Agro",
    "Consumer Cyclicals": "Consumo Masivo y Agro",
    "Basic Materials": "Construcción",
}


def map_sector(issuer, eco, bus, ind, gics, op_map):
    """Devuelve (sector, fuente). Prioridad:
       1) clasificación del operador para el MISMO emisor (consistencia),
       2) keywords del nombre (bancos/telecom/oil/gen/etc.),
       3) TRBC Economic Sector,
       4) None (a revisar)."""
    iss = (issuer or "").upper()
    norm = _norm_issuer(issuer)
    text = " ".join(x for x in (issuer, eco, bus, ind, gics) if x).upper()

    # 1) operador: match exacto o uno prefijo del otro (>=4 chars de raíz).
    if norm:
        if norm in op_map:
            return op_map[norm], "operator"
        for k, sec in op_map.items():
            if len(k) >= 3 and (norm.startswith(k + " ") or k.startswith(norm + " ")):
                return sec, "operator~"

    # 2) keywords. Bancos/financieras primero; consumer-agro ANTES que
    #    materiales (Profertil=fertilizante cae en Basic Materials por TRBC).
    fin = ("BANCO", "FINANCIERA", "CREDIT", "TARJETA", "MERCADO PAGO",
           "INDUSTRIAL AND COMMERCIAL BANK", "ICBC")
    if any(k in iss for k in fin) or "BANK" in iss:
        return "Sector Financiero", "issuer"
    if "TELECOM" in text:
        return "Telecomunicaciones", "issuer/trbc"
    cons = ("ARCOR", "SAN MIGUEL", "LEDESMA", "RIZOBACTER", "PROFERTIL",
            "JURAMENTO", "MUNOZ", "LIPSA", "HAVANNA", "FARMCITY",
            "FUTUROS Y OPCIONES", "MIRGOR")
    if any(k in iss for k in cons):
        return "Consumo Masivo y Agro", "issuer"
    if any(k in iss for k in ("AEROPUERTOS",)) or "AIRPORT" in text or "TRANSPORT" in text:
        return "Infraestructura y Transporte", "issuer/trbc"
    if any(k in iss for k in ("IRSA", "PLAZA LOGISTICA")) or "REAL ESTATE" in text:
        return "Bienes Raíces", "issuer/trbc"
    # generación / utilities ANTES que oil&gas (varias contienen "ENERGY")
    gen = ("ENERGIA ELECTRICA", "GREEN ENERGY", "GENNEIA", "360ENERGY",
           "CENTRAL PUERTO", "GENERACION", "LUZ DE TRES PICOS",
           "DISTRIBUIDORA DE ELECTRICIDAD")
    if any(k in iss for k in gen) or "UTILITIES" in text or "ELECTRIC UTIL" in text:
        return "Servicios Públicos y Generación", "issuer/trbc"
    oil = ("YPF", "VISTA", "PAN AMERICAN", "TECPETROL", "CROWN POINT", "CAPEX",
           "PECOM", "TANGO ENERGY", "COMBUSTIBLES", "PETROQUIMICA", "PETROLES",
           "OILTANKING")
    if any(k in iss for k in oil) or "ENERGY" in text or "OIL" in text or "GAS" in text or "PETRO" in text:
        return "Petróleo y Gas", "issuer/trbc"
    if any(k in text for k in ("CONSUMER", "FOOD", "AGRI", "RETAIL", "STAPLE")):
        return "Consumo Masivo y Agro", "trbc"

    # 3) TRBC económico
    if eco and eco in _TRBC_MAP:
        return _TRBC_MAP[eco], "trbc"
    return None, "unmapped"


def derive_amort(schedule, cur_tipo, vto):
    """(tipo, amort_inicio, amort_cant, amort_frec, amort_pct).

    tipo: respeta el del operador; si falta, lo infiere del schedule.
    BULLET: amort cantidad=1, amort inicio=vto (cosmético; el synth lo ignora).
    AMORTIZING: solo completa inicio/cantidad/frec/% si el schedule IPA es
    COMPLETO (suma de capital ≈ 100). Si ya amortizó parte (suma < 99.5), el
    schedule futuro es parcial -> se dejan en blanco (las flows exactas mandan).
    """
    caps = [c for c in (schedule or []) if c.get("amortization", 0) > 0.0001]
    cap_sum = sum(c["amortization"] for c in (schedule or []))

    if cur_tipo:
        tipo = cur_tipo.lower()
    elif schedule:
        tipo = "bullet" if len(caps) <= 1 else "amortizing"
    else:
        tipo = None

    if tipo == "bullet":
        return "bullet", vto, 1, None, None
    if tipo == "amortizing":
        if len(caps) < 2 or abs(cap_sum - 100) >= 0.5:
            return "amortizing", None, None, None, None  # parcial -> no inventar
        dates = [_to_date(c["date"]) for c in caps]
        gaps = [(dates[i + 1].year - dates[i].year) * 12 + (dates[i + 1].month - dates[i].month)
                for i in range(len(dates) - 1)]
        gaps = [g for g in gaps if g > 0]
        frec = None
        if gaps:
            f = round(12 / round(sum(gaps) / len(gaps)))
            frec = f if f in STD_FREQ else None
        amts = [round(c["amortization"], 4) for c in caps]
        pct = amts[0] if len(set(amts)) == 1 else None
        return "amortizing", dates[0], len(caps), frec, pct
    return None, None, None, None, None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--apply", action="store_true", help="escribe el Excel (sin esto, dry-run)")
    args = ap.parse_args()

    records = json.loads(FULL.read_text(encoding="utf-8"))
    op_map = load_operator_sector_map(XLSX)
    plan = []           # (ticker, fields_to_write)
    unmapped_sector = []
    unresolved = []     # ONs que LSEG no pudo resolver

    for rec in records:
        tk = rec["ticker"]
        cur = rec.get("current", {})
        ref = rec.get("ref") or {}
        sched = rec.get("schedule")
        if not ref.get("maturity") and not cur.get("fecha_vencimiento"):
            unresolved.append((tk, rec.get("status")))

        f = {}  # solo blancos

        def fill(col, val):
            if val is not None and not cur.get(col):
                f[col] = val

        fill("short_name", ref.get("issuer"))
        isin = cur.get("isin") or ref.get("isin")
        fill("isin", ref.get("isin"))
        fill("fecha_emision", _to_date(ref.get("issue_date")))
        fill("fecha_vencimiento", _to_date(ref.get("maturity")))
        if _num(ref.get("coupon_rate")) is not None:
            fill("cupon anual %", _num(ref.get("coupon_rate")))
        freq = ref.get("coupon_frequency")
        try:
            freq = int(freq) if freq not in (None, "") else None
        except (TypeError, ValueError):
            freq = None
        if freq in STD_FREQ:
            fill("frecuencia pagos", freq)

        leg = derive_legislacion(isin)
        fill("legislacion", leg)
        fill("base calculo", derive_base_calculo(leg or cur.get("legislacion")))

        sector, src = map_sector(ref.get("issuer") or cur.get("short_name"),
                                  ref.get("trbc_economic"), ref.get("trbc_business"),
                                  ref.get("trbc_industry"), ref.get("gics"), op_map)
        if not cur.get("sector"):
            if sector:
                f["sector"] = sector
            else:
                unmapped_sector.append((tk, ref.get("issuer") or cur.get("short_name"),
                                        ref.get("trbc_economic")))

        vto = _to_date(cur.get("fecha_vencimiento")) or _to_date(ref.get("maturity"))
        tipo, am_ini, am_cant, am_frec, am_pct = derive_amort(
            sched, cur.get("tipo amortizacion"), vto)
        fill("tipo amortizacion", tipo)
        fill("amort inicio", am_ini)
        fill("amort cantidad", am_cant)
        fill("amort frec", am_frec)
        if am_pct is not None:
            fill("amort %", am_pct)

        if f:
            f["ticker"] = tk
            plan.append((tk, f, sector if not cur.get("sector") else cur.get("sector")))

    print(f"{'APLICANDO' if args.apply else 'DRY-RUN'} · {len(plan)} ONs con campos a completar\n")
    for tk, f, sec in sorted(plan):
        cols = {k: v for k, v in f.items() if k != "ticker"}
        print(f"  {tk:<7} sector={sec or '???':<28} leg={f.get('legislacion','-'):<3} "
              f"base={f.get('base calculo','-'):<8} amort={f.get('tipo amortizacion','-'):<10} "
              f"+{sorted(cols)}")

    if unmapped_sector:
        print(f"\nSECTOR sin mapear ({len(unmapped_sector)}) — revisar/cargar a mano:")
        for tk, iss, eco in unmapped_sector:
            print(f"  {tk:<7} {iss} (TRBC={eco})")

    if unresolved:
        print(f"\nNO resueltas por LSEG ({len(unresolved)}):")
        for tk, st in unresolved:
            print(f"  {tk:<7} {st}")

    if args.apply:
        import openpyxl
        from apps.web.instruments_abm import (
            _LOCK, _write_instrument_row_on_wb, _atomic_save_workbook,
        )
        with _LOCK:
            wb = openpyxl.load_workbook(XLSX)
            try:
                for tk, f, _ in plan:
                    _write_instrument_row_on_wb(wb, "ONs", tk, f)
                _atomic_save_workbook(wb, XLSX)
            finally:
                wb.close()
        print(f"\nEscritas: {len(plan)} ONs (1 transacción atómica)")
    else:
        print(f"\nSe completarían: {len(plan)} ONs")


if __name__ == "__main__":
    main()
