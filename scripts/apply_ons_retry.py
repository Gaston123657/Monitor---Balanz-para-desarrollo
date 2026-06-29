"""Aplica las ONs resueltas de forma INEQUÍVOCA por lseg_retry_nomatch.py
(status=resolved_unique): completa campos + escribe cashflows exactos.

    py -3.12 scripts/apply_ons_retry.py            # dry-run
    py -3.12 scripts/apply_ons_retry.py --apply

Reusa las derivaciones de apply_ons_full.py (legislacion/base/sector/amort) y
el mismo escritor sancionado (instruments_abm).
"""
import argparse
import json
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from scripts.apply_ons_full import (  # noqa: E402
    XLSX, _to_date, _num, derive_legislacion, derive_base_calculo,
    map_sector, derive_amort, load_operator_sector_map, STD_FREQ,
)

RETRY = ROOT / "data" / "lseg_ons_retry.json"


def _cur(xlsx):
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    try:
        ws = wb["ONs"]; rows = list(ws.iter_rows(values_only=True))
        hdr = [str(c).strip().lower() if c else "" for c in rows[0]]
        ti = hdr.index("ticker")
        return {str(r[ti]).strip().upper(): {hdr[i]: r[i] for i in range(len(hdr))}
                for r in rows[1:] if r and r[ti]}
    finally:
        wb.close()


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--apply", action="store_true")
    ap.add_argument("--input", default=str(RETRY),
                    help="JSON con registros resolved_unique (default: lseg_ons_retry.json)")
    args = ap.parse_args()

    recs = json.loads(Path(args.input).read_text(encoding="utf-8"))
    op_map = load_operator_sector_map(XLSX)
    cur = _cur(XLSX)
    plan = []
    for rec in recs:
        if rec.get("status") != "resolved_unique":
            continue
        tk = rec["ticker"]; ref = rec["ref"]; sched = rec.get("schedule")
        c = cur.get(tk, {})
        f = {"ticker": tk}

        def fill(col, val):
            if val is not None and not c.get(col):
                f[col] = val

        fill("short_name", ref.get("issuer"))
        isin = c.get("isin") or ref.get("isin")
        fill("isin", ref.get("isin"))
        fill("fecha_emision", _to_date(ref.get("issue_date")))
        fill("fecha_vencimiento", _to_date(ref.get("maturity")))
        if _num(ref.get("coupon_rate")) is not None:
            fill("cupon anual %", _num(ref.get("coupon_rate")))
        freq = ref.get("coupon_frequency")
        try:
            freq = int(freq) if freq not in (None, "") else None
        except (TypeError, ValueError):
            freq = None
        fill("frecuencia pagos", freq if freq in STD_FREQ else None)
        leg = derive_legislacion(isin)
        fill("legislacion", leg)
        fill("base calculo", derive_base_calculo(leg or c.get("legislacion")))
        sector, _ = map_sector(ref.get("issuer"), ref.get("trbc_economic"),
                               ref.get("trbc_business"), ref.get("trbc_industry"),
                               ref.get("gics"), op_map)
        if sector and not c.get("sector"):
            f["sector"] = sector
        vto = _to_date(c.get("fecha_vencimiento")) or _to_date(ref.get("maturity"))
        tipo, ai, ac, af, ap_ = derive_amort(sched, c.get("tipo amortizacion"), vto)
        fill("tipo amortizacion", tipo)
        fill("amort inicio", ai); fill("amort cantidad", ac); fill("amort frec", af)
        if ap_ is not None:
            fill("amort %", ap_)

        cfs = None
        if sched and not str(ref.get("coupon_type") or "").startswith("FR"):
            cfs = [{"date": x["date"], "amortization": x["amortization"],
                    "interest": x["interest"]} for x in sched]
        plan.append((tk, f, cfs))

    print(f"{'APLICANDO' if args.apply else 'DRY-RUN'} · {len(plan)} ONs únicas\n")
    for tk, f, cfs in plan:
        print(f"  {tk:<7} {{k:v}}={ {k: v for k, v in f.items() if k != 'ticker'} } cfs={len(cfs) if cfs else 0}")

    if args.apply:
        from apps.web.instruments_abm import (
            _LOCK, _write_instrument_row_on_wb, _write_cashflows_on_wb,
            _parse_cashflows, _atomic_save_workbook,
        )
        with _LOCK:
            wb = openpyxl.load_workbook(XLSX)
            try:
                for tk, f, cfs in plan:
                    _write_instrument_row_on_wb(wb, "ONs", tk, f)
                    if cfs is not None:
                        _write_cashflows_on_wb(wb, tk.upper(), _parse_cashflows(cfs))
                _atomic_save_workbook(wb, XLSX)
            finally:
                wb.close()
        print(f"\nEscritas: {len(plan)} ONs")


if __name__ == "__main__":
    main()
