"""One-shot: migrate ONs.xlsx (standalone project) into instruments_master.xlsx as a new sheet.

Reads:  C:\\Users\\glatigano\\monitor\\ONs\\ONs.xlsx        (164 rows, single sheet)
Writes: data/instruments_master.xlsx                       (new sheet "ONs")

Only the MEP ticker (suffix D) is migrated — the panel filters to D only,
mirroring how bonares/bopreales render. Pesos (O) and Cable (C) tickers stay
in the standalone file for now; the corp endpoint /1000 normalization in
repositories.py covers their pricing if/when we add a popup later.

Idempotent: if the "ONs" sheet already exists in master, it is overwritten.
"""

from __future__ import annotations

import os
import sys
from pathlib import Path

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO))

import openpyxl
from openpyxl.styles import PatternFill
from apps.web.instruments_abm import _atomic_save_workbook

SRC = Path(r"C:\Users\glatigano\monitor\ONs\ONs.xlsx")
DST = REPO / "data" / "instruments_master.xlsx"

# (target_header, source_header_or_None_for_constant)
COL_MAP = [
    ("ticker",             "Ticker MEP"),
    ("short_name",         "Nombre"),
    ("tipo",               None),               # constant "ON"
    ("sector",             "Sector"),
    ("legislacion",        "Legislacion"),
    ("isin",               "ISIN"),
    ("fecha_emision",      "Fecha Emision"),
    ("fecha_vencimiento",  "Fecha Vto"),
    ("cupon anual %",      "Cupon %"),
    ("frecuencia pagos",   "Frecuencia"),
    ("base calculo",       "Base"),
    ("tipo amortizacion",  "Tipo Amort"),
    ("amort inicio",       "Amort Inicio"),
    ("amort cantidad",     "Amort Cant"),
    ("amort frec",         "Amort Frec"),
    ("amort %",            "Amort %"),
    ("amort cant1",        "Amort Cant1"),
    ("amort %2",           "Amort %2"),
]

YELLOW = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Soft yellow for ON header


def _normalize_legislacion(v):
    if v is None:
        return None
    s = str(v).strip().upper()
    return "AR" if s == "ARG" else s  # "NY" stays "NY"


def _normalize_base(v):
    """Map "real/365" → "ACT/365" so the engine recognizes it. Pass through otherwise."""
    if v is None:
        return None
    s = str(v).strip()
    if s.lower() in ("real/365", "actual/365"):
        return "ACT/365"
    return s


def _normalize_tipo_amort(v):
    if v is None:
        return None
    return str(v).strip().lower()


def main():
    if not SRC.exists():
        raise SystemExit(f"Source not found: {SRC}")
    if not DST.exists():
        raise SystemExit(f"Master not found: {DST}")

    print(f"Reading {SRC}...")
    wb_src = openpyxl.load_workbook(SRC, data_only=True, read_only=True)
    ws_src = wb_src[wb_src.sheetnames[0]]
    rows_src = list(ws_src.iter_rows(values_only=True))
    headers_src = [str(h).strip() if h else "" for h in rows_src[0]]
    idx = {h: i for i, h in enumerate(headers_src)}
    print(f"Source rows: {len(rows_src) - 1}")

    print(f"Loading master {DST}...")
    wb = openpyxl.load_workbook(DST)
    if "ONs" in wb.sheetnames:
        print('Existing "ONs" sheet found — removing.')
        del wb["ONs"]
    ws = wb.create_sheet("ONs")

    target_headers = [h for h, _ in COL_MAP]
    ws.append(target_headers)
    for c, cell in enumerate(ws[1], start=1):
        cell.fill = YELLOW

    migrated = skipped = 0
    skipped_no_mep = 0
    for src_row in rows_src[1:]:
        if not src_row or src_row[idx["Ticker MEP"]] is None:
            skipped_no_mep += 1
            continue
        out = []
        for target_h, src_h in COL_MAP:
            if src_h is None:
                out.append("ON")
                continue
            v = src_row[idx[src_h]] if src_h in idx else None
            if target_h == "legislacion":
                v = _normalize_legislacion(v)
            elif target_h == "base calculo":
                v = _normalize_base(v)
            elif target_h == "tipo amortizacion":
                v = _normalize_tipo_amort(v)
            out.append(v)
        ws.append(out)
        migrated += 1

    _atomic_save_workbook(wb, str(DST))
    wb.close()
    wb_src.close()

    print(f"Migrated:   {migrated} rows into sheet 'ONs'")
    print(f"Skipped:    {skipped_no_mep} (no MEP ticker)")
    print(f"Done — master saved atomically.")


if __name__ == "__main__":
    main()
