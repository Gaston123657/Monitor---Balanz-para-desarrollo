"""Segundo pase: completa las ONs que quedaron parciales tras apply_ons_full.py.

    py -3.12 scripts/apply_ons_partials.py            # dry-run
    py -3.12 scripts/apply_ons_partials.py --apply

Casos (todos con ref de LSEG, pero freq/cupón/tipo no estándar):
  - Zero-coupon (FXZC) y bullets de pago único (FXPM): cupón ya está, falta
    frecuencia -> se infiere del schedule (sin cupones -> 1).
  - Amortizing irregular (Rizobacter BIOX, freq=777 de LSEG): se infiere la
    frecuencia del cadence de intereses y se ESCRIBEN las flows exactas de LSEG
    (el synth no reproduce un schedule irregular).
  - FRN en ARS (FRPV): cupón variable (BADLAR) -> el monitor no las puede
    pricear con flujo fijo; solo se setea tipo=bullet. Quedan sin cupón fijo.

Para cada ticker con schedule de LSEG, escribe además las flows EXACTAS a la
hoja Cashflows (estas ONs son nuevas y no las tenían). Todo vía instruments_abm
en una transacción atómica.
"""
import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

XLSX = str(ROOT / "data" / "instruments_master.xlsx")
FULL = ROOT / "data" / "lseg_ons_full.json"
STD_FREQ = {1, 2, 4, 12}

# Tickers parciales con datos de LSEG (excluye los 16 no_match y YMCZD ref_error).
TARGETS = ["LUC4D", "OLC3D", "MGCED", "YMCTD", "AFCHD", "AFCJD", "SXC4D",
           "VSCID", "RZBBD", "RZBCD", "RZBAD", "MR39D", "RC3CD", "RC6CD"]


def _to_date(iso):
    return datetime.strptime(str(iso)[:10], "%Y-%m-%d").date() if iso else None


def _months(a, b):
    return (b.year - a.year) * 12 + (b.month - a.month)


def infer_freq(schedule):
    """Frecuencia de cupón del cadence de flujos con interés. ZC -> 1."""
    if not schedule:
        return 1
    idates = [_to_date(c["date"]) for c in schedule if c.get("interest", 0) > 0.0001]
    if len(idates) <= 1:
        return 1
    gaps = [_months(idates[i], idates[i + 1]) for i in range(len(idates) - 1)]
    gaps = [g for g in gaps if g > 0]
    if not gaps:
        return 1
    f = round(12 / (sum(gaps) / len(gaps)))
    return f if f in STD_FREQ else 2


def _cur_values(xlsx):
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    try:
        ws = wb["ONs"]
        rows = list(ws.iter_rows(values_only=True))
        hdr = [str(c).strip().lower() if c else "" for c in rows[0]]
        ti = hdr.index("ticker")
        out = {}
        for r in rows[1:]:
            if r and r[ti]:
                out[str(r[ti]).strip().upper()] = {
                    hdr[i]: r[i] for i in range(len(hdr))}
        return out
    finally:
        wb.close()


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--apply", action="store_true")
    args = ap.parse_args()

    recs = {r["ticker"]: r for r in json.loads(FULL.read_text(encoding="utf-8"))}
    cur = _cur_values(XLSX)
    plan = []  # (ticker, fields, cashflows|None)

    for tk in TARGETS:
        rec = recs.get(tk, {})
        ref = rec.get("ref") or {}
        sched = rec.get("schedule") or []
        c = cur.get(tk, {})
        f = {"ticker": tk}

        caps = [x for x in sched if x.get("amortization", 0) > 0.0001]
        is_floating = str(ref.get("coupon_type") or "").startswith("FR")

        # frecuencia (solo si falta)
        if not c.get("frecuencia pagos"):
            f["frecuencia pagos"] = ref.get("coupon_frequency") if ref.get("coupon_frequency") in STD_FREQ else infer_freq(sched)
        # tipo amortizacion (solo si falta)
        if not c.get("tipo amortizacion"):
            f["tipo amortizacion"] = "amortizing" if len(caps) > 1 else "bullet"

        # cashflows exactos de LSEG (solo si hay schedule fijo; FRN no tiene)
        cfs = None
        if sched and not is_floating:
            cfs = [{"date": x["date"], "amortization": x["amortization"],
                    "interest": x["interest"]} for x in sched]

        if len(f) > 1 or cfs is not None:
            plan.append((tk, f, cfs))

    print(f"{'APLICANDO' if args.apply else 'DRY-RUN'} · {len(plan)} ONs parciales\n")
    for tk, f, cfs in plan:
        extra = {k: v for k, v in f.items() if k != "ticker"}
        print(f"  {tk:<7} {extra}  cashflows={len(cfs) if cfs else 0}")

    if args.apply:
        from apps.web.instruments_abm import (
            _LOCK, _write_instrument_row_on_wb, _write_cashflows_on_wb,
            _parse_cashflows, _atomic_save_workbook,
        )
        with _LOCK:
            wb = openpyxl.load_workbook(XLSX)
            try:
                for tk, f, cfs in plan:
                    _write_instrument_row_on_wb(wb, "ONs", tk, f)
                    if cfs is not None:
                        _write_cashflows_on_wb(wb, tk.upper(), _parse_cashflows(cfs))
                _atomic_save_workbook(wb, XLSX)
            finally:
                wb.close()
        print(f"\nEscritas: {len(plan)} ONs (1 transacción atómica)")


if __name__ == "__main__":
    main()
