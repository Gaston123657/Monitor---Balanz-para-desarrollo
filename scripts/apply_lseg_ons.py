"""Aplica al master Excel los datos de ONs bajados de LSEG (cache JSON).

Corre con el Python del MONITOR (py 3.12), NO con el venv aislado:

    py -3.12 scripts/apply_lseg_ons.py            # dry-run (no escribe)
    py -3.12 scripts/apply_lseg_ons.py --apply    # escribe el Excel

Lee data/lseg_ons_cache.json (generado por scripts/lseg_fetch_ons.py en el
venv .venv-lseg) y, para cada ON con status="ok", completa en la hoja `ONs`
del master: vencimiento, cupón, frecuencia, emisión, ISIN y el SCHEDULE de
cashflows exacto (hoja Cashflows). Así el calendario de pagos las estima sin
volver a depender de la API.

Escribe SIEMPRE vía instruments_abm.save_instrument (único escritor sancionado
del Excel: atomic .tmp + os.replace, mismo _LOCK que el repo). NO toca
sector/legislacion existentes (se preservan). Las ONs no resueltas
(no_match / floating / ambiguous) se listan al final para carga manual.
"""
import argparse
import json
import sys
from datetime import date, datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from apps.web import instruments_abm

XLSX = str(ROOT / "data" / "instruments_master.xlsx")
CACHE = ROOT / "data" / "lseg_ons_cache.json"
STD_FREQ = {1, 2, 4, 12}


def _to_date(iso):
    if not iso:
        return None
    return datetime.strptime(str(iso)[:10], "%Y-%m-%d").date()


def _num(v):
    if v in (None, ""):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _existing_short_names():
    """{ticker: short_name} actuales en la hoja ONs (para no pisar nombres)."""
    import openpyxl
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    try:
        ws = wb["ONs"]
        rows = list(ws.iter_rows(values_only=True))
        hdr = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
        ti, si = hdr.index("ticker"), hdr.index("short_name")
        out = {}
        for r in rows[1:]:
            if r and r[ti]:
                nm = r[si]
                nm = str(nm).strip() if nm is not None else ""
                out[str(r[ti]).strip().upper()] = "" if nm.lower() in ("nan", "none") else nm
        return out
    finally:
        wb.close()


def build(rec, existing_name):
    """Devuelve (fields, cashflows) listos para save_instrument."""
    ref = rec["ref"]
    sched = rec["schedule"]

    freq = ref.get("coupon_frequency")
    try:
        freq = int(freq) if freq not in (None, "") else None
    except (TypeError, ValueError):
        freq = None
    if freq not in STD_FREQ:
        freq = None  # 777/0/'' -> en blanco; el repo lo infiere de los cashflows

    cap_flows = [c for c in sched if c["amortization"] > 0]
    cap_sum = sum(c["amortization"] for c in sched)
    amort = "amortizing" if (len(cap_flows) > 1 or cap_sum < 99.5) else "bullet"

    fields = {
        "ticker": rec["ticker"],
        "tipo": "ON",
        "fecha_emision": _to_date(ref.get("issue_date")),
        "fecha_vencimiento": _to_date(ref.get("maturity")),
        "cupon anual %": _num(ref.get("coupon_rate")) or 0.0,
        "frecuencia pagos": freq,
        "tipo amortizacion": amort,
    }
    isin = ref.get("isin") or rec.get("isin_local")
    if isin:
        fields["isin"] = isin
    if not existing_name and ref.get("issuer"):
        fields["short_name"] = ref["issuer"]

    cashflows = [{"date": c["date"],
                  "amortization": c["amortization"],
                  "interest": c["interest"]} for c in sched]
    return fields, cashflows


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--apply", action="store_true", help="escribe el Excel (sin esto, dry-run)")
    args = ap.parse_args()

    records = json.loads(CACHE.read_text(encoding="utf-8"))
    ok = [r for r in records if r.get("status") == "ok"]
    other = [r for r in records if r.get("status") != "ok"]
    names = _existing_short_names()

    print(f"{'APLICANDO' if args.apply else 'DRY-RUN'} · {len(ok)} ONs ok · cache {CACHE.name}\n")
    plan = []
    for rec in ok:
        tk = rec["ticker"]
        fields, cfs = build(rec, names.get(tk, ""))
        plan.append((tk, fields, cfs))
        print(f"  {tk:<6} vto={fields['fecha_vencimiento']} cpn={fields['cupon anual %']:<6} "
              f"freq={fields['frecuencia pagos']} {fields['tipo amortizacion']:<10} "
              f"cfs={len(cfs)} isin={fields.get('isin','-')}")

    if args.apply:
        # UNA sola transacción: load -> aplicar todo en memoria -> un atomic save.
        # Evita el churn de 108 os.replace seguidos sobre un Excel en OneDrive
        # (gatilla PermissionError intermitente) y es all-or-nothing.
        import openpyxl
        from apps.web.instruments_abm import (
            _LOCK, _write_instrument_row_on_wb, _write_cashflows_on_wb,
            _parse_cashflows, _atomic_save_workbook,
        )
        with _LOCK:
            wb = openpyxl.load_workbook(XLSX)
            try:
                for tk, fields, cfs in plan:
                    _write_instrument_row_on_wb(wb, "ONs", tk, fields)
                    _write_cashflows_on_wb(wb, tk.upper(), _parse_cashflows(cfs))
                _atomic_save_workbook(wb, XLSX)
            finally:
                wb.close()
        print(f"\nEscritas: {len(plan)} ONs (1 transacción atómica)")
    else:
        print(f"\nSe escribirían: {len(plan)} ONs")

    if other:
        print(f"\nNO aplicadas ({len(other)}) — requieren carga manual:")
        from collections import defaultdict
        by = defaultdict(list)
        for r in other:
            by[r["status"]].append(r["ticker"])
        for st, tks in sorted(by.items()):
            print(f"  {st:<12}: {', '.join(sorted(tks))}")


if __name__ == "__main__":
    main()
