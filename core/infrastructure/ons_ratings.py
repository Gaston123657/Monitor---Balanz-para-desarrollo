"""Calificaciones crediticias de largo plazo FIXscr (FIX SCR S.A., afiliada Fitch).

Fuente: export de FIXscr en ``data/fixscr_ratings.xlsx`` (columnas ENTIDAD, FECHA,
PAÍS, AREA, SECTOR, TIPO DE CALIFICACIÓN, CALIFICACIÓN CORTO PLAZO, CALIFICACIÓN
LARGO PLAZO, PERSPECTIVA / RATING WATCH, ESTADO). Para actualizar los ratings basta
con reemplazar ese archivo por un export nuevo — este módulo lo recarga al detectar
un cambio de mtime; sólo emisores nuevos requieren una línea en ``_ALIASES``.

El match es **a nivel emisor** (las filas por clase del export traen el nombre de la
clase truncado con "...", no son confiables para mapear a un ticker puntual). Por
cada emisor FIXscr se toma su rating de largo plazo a nivel emisor (filas tipo
"Emisor" / "Endeudamiento de Largo Plazo" / "Calificación de largo plazo" /
"Fortaleza Financiera de Largo Plazo"); si no tiene fila a nivel emisor, la moda de
las calificaciones LP de sus filas por instrumento.

El matcheo de nuestros ``short_name`` contra las ENTIDAD de FIXscr es deterministco:
1. ``_ALIASES`` (mapa curado: correcciones, abreviaturas y blocklist de emisores que
   FIXscr NO califica) — máxima prioridad para evitar falsos positivos.
2. Igualdad exacta del nombre normalizado.
3. Substring de palabra completa (con guarda de longitud) — seguro contra near-misses
   tipo TGS↔TGN.
Lo que no resuelve queda como ``None`` (s/c — sin calificación FIXscr), nunca un
match inventado.
"""
from __future__ import annotations

import logging
import re
import threading
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path
from typing import Dict, Optional

logger = logging.getLogger(__name__)

_XLSX_PATH = Path(__file__).resolve().parents[2] / "data" / "fixscr_ratings.xlsx"

# Filas cuyo TIPO DE CALIFICACIÓN representa el rating de largo plazo del EMISOR.
_ISSUER_TYPES = {
    "Emisor",
    "Endeudamiento de Largo Plazo",
    "Calificación de largo plazo",
    "Fortaleza Financiera de Largo Plazo",
}

# Tokens de forma jurídica / genéricos que se descartan al normalizar nombres.
_STOP = {
    "sa", "sau", "saic", "saicf", "sacif", "sacifia", "srl", "sl", "saai", "saci",
    "ci", "cia", "ciasa", "sas", "sgr", "aici", "agici", "aif", "de", "del", "la",
    "el", "y", "e", "f",
}

# Mapa curado our-short_name(substring, lower, sin acentos) -> búsqueda FIXscr.
#   None  => el emisor NO está calificado por FIXscr (s/c) — bloquea falsos positivos.
#   str   => substring a buscar dentro del nombre normalizado de la ENTIDAD FIXscr.
# El orden importa: se evalúa secuencialmente y gana el primer hit.
_ALIASES = [
    # --- Blocklist: FIXscr no los califica (rated por otra agencia o sin rating) ---
    ("transportadora de gas del sur", None),   # sólo existe TGN en FIXscr, NO TGS
    ("mastellone", None),
    ("industrial and commercial", None),       # ICBC banco (sólo está la admin. de FCI)
    ("genneia", None),
    ("mirgor", None),
    ("supervielle", None),
    ("clisa", None),
    ("impsa", None),
    ("farmcity", None),
    ("oiltanking", None),
    ("ebytem", None),
    ("tarjeta naranja", None),
    ("petroles sudamericanos", None),
    ("generacion litoral", None),
    ("banco patagonia", None),                 # sólo está Patagonia Inversora (FCI)
    ("banco mariva", None),                    # sólo está Mariva Asset Management (FCI)
    # --- Abreviaturas / variantes / nombres parciales ---
    ("telecom", "telecom argentina"),
    ("irsa", "irsa inversiones"),
    ("pan american", "pan american energy"),
    ("cgc", "general de combustibles"),
    ("compania general de combustibles", "general de combustibles"),
    ("cnh", "cnh industrial"),
    ("ledesma", "ledesma"),
    ("meranol", "meranol"),
    ("360", "360 energy"),
    ("scc power", "scc power san pedro"),
    ("banco galicia", "galicia y buenos aires"),
    ("comercializadora norte", "edenor"),
    ("edenor", "edenor"),
    ("electricidad de salta", "electricidad de salta"),
    ("electricidad de mendoza", "electricidad de mendoza"),
    ("aluar", "aluar"),
    ("havanna", "havanna"),
]


def _strip_accents(s) -> str:
    return unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode()


def _norm(s) -> str:
    """Normaliza un nombre: sin acentos, minúsculas, sin puntos/puntuación, sin
    formas jurídicas ni tokens de 1 letra. "YPF S.A." y "YPF SA" -> "ypf"."""
    s = _strip_accents(s).lower().replace(".", "")
    s = re.sub(r"[^a-z0-9 ]", " ", s)
    return " ".join(t for t in s.split() if len(t) > 1 and t not in _STOP).strip()


# --------------------------------------------------------------------------- #
# Carga con cache por mtime (recarga automática al reemplazar el xlsx).
# --------------------------------------------------------------------------- #
_lock = threading.Lock()
_cache_mtime: Optional[float] = None
_fix_index: list = []          # [(norm_entidad, entidad, rating, perspectiva, fecha)]
_short_cache: Dict[str, Optional[dict]] = {}


def _load() -> None:
    """(Re)carga el índice FIXscr si el archivo cambió. Idempotente y barato."""
    global _cache_mtime, _fix_index, _short_cache
    try:
        mtime = _XLSX_PATH.stat().st_mtime
    except OSError:
        logger.warning("ons_ratings: %s no encontrado — ratings deshabilitados", _XLSX_PATH)
        _fix_index = []
        _short_cache = {}
        _cache_mtime = None
        return
    if _cache_mtime == mtime and _fix_index:
        return
    try:
        import openpyxl
        wb = openpyxl.load_workbook(_XLSX_PATH, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))[1:]
    except Exception:
        logger.exception("ons_ratings: fallo al leer %s", _XLSX_PATH)
        return

    issuer_lvl: Dict[str, tuple] = {}            # entidad -> (rating, perspectiva, fecha)
    on_lp: Dict[str, Counter] = defaultdict(Counter)
    for r in rows:
        ent = r[0]
        if not ent:
            continue
        ent = str(ent).strip()
        tipo = str(r[5] or "")
        lp = r[7]
        if tipo in _ISSUER_TYPES and lp and ent not in issuer_lvl:
            issuer_lvl[ent] = (str(lp), str(r[8] or "").strip(), _fmt_date(r[1]))
        if isinstance(lp, str) and lp.endswith("(arg)") and "f(" not in lp and "c(" not in lp:
            on_lp[ent][lp] += 1

    def rating_of(ent: str):
        if ent in issuer_lvl:
            return issuer_lvl[ent]
        if on_lp.get(ent):
            return (on_lp[ent].most_common(1)[0][0], "", "")
        return None

    index = []
    for ent in set([*issuer_lvl, *on_lp]):
        rt = rating_of(ent)
        if rt:
            index.append((_norm(ent), ent, rt[0], rt[1], rt[2]))

    with _lock:
        _fix_index = index
        _short_cache = {}
        _cache_mtime = mtime
    logger.info("ons_ratings: %d entidades FIXscr con rating LP cargadas", len(index))


def _fmt_date(v) -> str:
    if v is None:
        return ""
    s = str(v)
    return s[:10]


def _find_contains(search: str) -> Optional[tuple]:
    """Primera entidad FIXscr cuyo nombre normalizado == o contiene `search`."""
    sn = _norm(search)
    if not sn:
        return None
    pat = re.compile(r"\b" + re.escape(sn) + r"\b")
    for fn, ent, rating, persp, fecha in _fix_index:
        if sn == fn or pat.search(fn):
            return (ent, rating, persp, fecha)
    return None


def _resolve(short_name: str) -> Optional[dict]:
    if not short_name:
        return None
    low = _strip_accents(short_name).lower()
    # 1. alias curado (correcciones + blocklist)
    for key, target in _ALIASES:
        if key in low:
            if target is None:
                return None
            hit = _find_contains(target)
            return _pack(hit) if hit else None
    n = _norm(short_name)
    if not n:
        return None
    # 2. igualdad exacta normalizada
    for fn, ent, rating, persp, fecha in _fix_index:
        if fn == n:
            return _pack((ent, rating, persp, fecha))
    # 3. substring de palabra completa, con guarda (>=2 tokens y >=6 chars)
    for fn, ent, rating, persp, fecha in _fix_index:
        a, b = (n, fn) if len(n) <= len(fn) else (fn, n)
        if len(a) >= 6 and len(a.split()) >= 2 and re.search(r"\b" + re.escape(a) + r"\b", b):
            return _pack((ent, rating, persp, fecha))
    return None


def _pack(hit: tuple) -> dict:
    ent, rating, persp, fecha = hit
    return {
        "rating": rating,
        "perspectiva": persp or None,
        "fecha": fecha or None,
        "entity": ent,
        "source": "FIXscr",
    }


def get_rating(short_name: Optional[str]) -> Optional[dict]:
    """Calificación FIXscr de largo plazo para el emisor de un short_name.

    Devuelve dict {rating, perspectiva, fecha, entity, source} o ``None`` si el
    emisor no está calificado por FIXscr (s/c). Thread-safe y cacheado por
    short_name.
    """
    if not short_name:
        return None
    _load()
    key = str(short_name).strip()
    if key in _short_cache:
        return _short_cache[key]
    res = _resolve(key)
    with _lock:
        _short_cache[key] = res
    return res
