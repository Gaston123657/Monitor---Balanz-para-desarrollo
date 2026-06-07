"""Background sync de Obligaciones Negociables nuevas detectadas en data912.com/live/arg_corp.

Reemplaza al `ONs/updater.py` del proyecto standalone. Cada vez que arranca
el server (y cada 24h después), descubre tickers base (suffix-stripped) que
están en arg_corp pero no en la hoja `ONs` del master Excel y los inserta
con fondo amarillo + campos no-críticos vacíos. Operador completa después
vía la ABM.

Sin esto, agregar una ON nueva requería:
  1. ver que apareció en monitor.py legacy,
  2. correr updater.py a mano,
  3. completar fields a mano.
Ahora basta con esperar al próximo ciclo y completar vía la ABM web.
"""

from __future__ import annotations

import logging
import threading
import time
from pathlib import Path
from typing import Iterable, Set

import openpyxl
from openpyxl.styles import PatternFill

from apps.web.instruments_abm import _LOCK as _ABM_LOCK, _atomic_save_workbook
from core.infrastructure._http import http_get_json

logger = logging.getLogger(__name__)

API_URL = "https://data912.com/live/arg_corp"
SHEET = "ONs"
SYNC_PERIOD_SEC = 24 * 3600  # 1×/día — el universo de ONs cambia raramente
YELLOW = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

# Headers de la hoja ONs (matchean el schema de instruments_abm.SHEET_SCHEMAS["ONs"]).
_ONS_HEADERS = (
    "ticker", "short_name", "tipo", "sector", "legislacion", "isin",
    "fecha_emision", "fecha_vencimiento", "cupon anual %", "frecuencia pagos",
    "base calculo", "tipo amortizacion", "amort inicio", "amort cantidad",
    "amort frec", "amort %", "amort cant1", "amort %2",
)

# Estado del último sync — leído por /api/snapshot para alimentar el cartel
# "N ONs nuevas detectadas" en el dashboard. ts es unix epoch para que el
# frontend pueda comparar contra el timestamp de dismiss en localStorage.
_LAST_RESULT_LOCK = threading.Lock()
_LAST_RESULT: dict | None = None


def get_last_sync_result() -> dict | None:
    """Resultado del último `sync_new_ons()`, o None si todavía no corrió.

    Shape: {"added": int, "checked": int, "skipped_excluded": int,
            "ts": float, "new_tickers": [str, ...], "error": str?}
    """
    with _LAST_RESULT_LOCK:
        return dict(_LAST_RESULT) if _LAST_RESULT else None


def _set_last_result(result: dict) -> None:
    with _LAST_RESULT_LOCK:
        global _LAST_RESULT
        _LAST_RESULT = dict(result)


def _load_exclude(exclude_path: Path) -> Set[str]:
    if not exclude_path.exists():
        return set()
    out: Set[str] = set()
    for line in exclude_path.read_text(encoding="utf-8").splitlines():
        s = line.strip()
        if s and not s.startswith("#"):
            out.add(s.upper())
    return out


def _fetch_corp_symbols() -> Iterable[str]:
    """Devuelve la lista de tickers cotizando en arg_corp."""
    payload = http_get_json(API_URL, timeout=10, source="ons_sync/arg_corp")
    if not isinstance(payload, list):
        raise ValueError(f"arg_corp expected list, got {type(payload).__name__}")
    return [str(item.get("symbol", "")).upper() for item in payload if item.get("symbol")]


def _group_by_base(symbols: Iterable[str]) -> dict:
    """{'PNDC': {'O': 'PNDCO', 'D': 'PNDCD', 'C': 'PNDCC'}, ...}"""
    groups: dict = {}
    for sym in symbols:
        if len(sym) < 2:
            continue
        suffix = sym[-1]
        if suffix not in ("O", "D", "C"):
            continue
        base = sym[:-1]
        groups.setdefault(base, {})[suffix] = sym
    return groups


def _existing_bases(ws) -> Set[str]:
    """Bases conocidas en la hoja ONs (suffix-stripped del ticker MEP)."""
    headers = [str(c.value).lower().strip() if c.value else "" for c in ws[1]]
    if "ticker" not in headers:
        return set()
    tcol = headers.index("ticker") + 1
    out: Set[str] = set()
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=tcol).value
        if v is None:
            continue
        s = str(v).strip().upper()
        if len(s) > 1:
            out.add(s[:-1])  # strip suffix (D/O/C)
    return out


def sync_new_ons(xlsx_path: str, exclude_path: str) -> dict:
    """Detecta ONs nuevas en arg_corp y las agrega a la hoja ONs del master.

    Solo inserta filas con `ticker = MEP (sufijo D)`, `tipo = "ON"`. Resto
    de los campos quedan vacíos para que el operador los complete vía ABM.
    Marca las filas nuevas con fondo amarillo para visibilidad.

    Idempotente y thread-safe: comparte `_ABM_LOCK` con el resto de las
    operaciones I/O sobre el master.

    Returns: {"checked": int, "added": int, "skipped_excluded": int}
    """
    try:
        symbols = list(_fetch_corp_symbols())
    except Exception as e:
        logger.warning(f"ons_sync: arg_corp fetch failed: {e}")
        result = {"checked": 0, "added": 0, "skipped_excluded": 0,
                  "ts": time.time(), "new_tickers": [], "error": str(e)}
        _set_last_result(result)
        return result

    groups = _group_by_base(symbols)
    excluded = _load_exclude(Path(exclude_path))

    with _ABM_LOCK:
        wb = openpyxl.load_workbook(xlsx_path)
        try:
            if SHEET not in wb.sheetnames:
                ws = wb.create_sheet(SHEET)
                ws.append(list(_ONS_HEADERS))
            else:
                ws = wb[SHEET]

            known = _existing_bases(ws)
            new_bases = sorted(b for b in groups if b not in known and b not in excluded)
            skipped_excluded = sum(1 for b in groups if b in excluded and b not in known)

            if not new_bases:
                logger.info(
                    f"ons_sync: {len(groups)} bases cotizando, "
                    f"{len(known)} ya cargadas, 0 nuevas."
                )
                result = {"checked": len(groups), "added": 0,
                          "skipped_excluded": skipped_excluded,
                          "ts": time.time(), "new_tickers": []}
                _set_last_result(result)
                return result

            header_row = [str(c.value).lower().strip() if c.value else "" for c in ws[1]]
            ticker_idx = header_row.index("ticker") if "ticker" in header_row else 0
            tipo_idx = header_row.index("tipo") if "tipo" in header_row else 2

            first_new_row = ws.max_row + 1
            new_tickers: list[str] = []
            for base in new_bases:
                tickers = groups[base]
                mep = tickers.get("D")
                if not mep:
                    # Sin MEP no hay nada que monitorear; saltamos hasta tener
                    # data912 publicando el ticker D.
                    continue
                row_vals = [None] * len(header_row)
                row_vals[ticker_idx] = mep
                row_vals[tipo_idx] = "ON"
                ws.append(row_vals)
                new_tickers.append(mep)

            # Pintar amarillo las filas insertadas para que el operador las vea.
            for r in range(first_new_row, ws.max_row + 1):
                for c in range(1, len(header_row) + 1):
                    ws.cell(row=r, column=c).fill = YELLOW

            _atomic_save_workbook(wb, xlsx_path)
            added = ws.max_row + 1 - first_new_row
            logger.info(
                f"ons_sync: {len(groups)} bases, +{added} nuevas → {new_bases[:10]}"
                f"{'…' if added > 10 else ''}"
            )
            result = {"checked": len(groups), "added": added,
                      "skipped_excluded": skipped_excluded,
                      "ts": time.time(), "new_tickers": new_tickers}
            _set_last_result(result)
            return result
        finally:
            wb.close()


def start_background_sync(xlsx_path: str, exclude_path: str,
                          shutdown_event: "threading.Event") -> threading.Thread:
    """Spawn daemon thread: sync_new_ons() al boot y cada 24h después.

    Stops cuando shutdown_event se setea (no en mid-loop, sino al despertar
    de su próximo sleep).
    """
    def _loop():
        # Boot: corremos enseguida pero en background — no bloqueamos el listen.
        try:
            sync_new_ons(xlsx_path, exclude_path)
        except Exception:
            logger.exception("ons_sync boot run failed")
        while not shutdown_event.wait(SYNC_PERIOD_SEC):
            try:
                sync_new_ons(xlsx_path, exclude_path)
            except Exception:
                logger.exception("ons_sync periodic run failed")

    t = threading.Thread(target=_loop, daemon=True, name="ons_sync")
    t.start()
    return t
